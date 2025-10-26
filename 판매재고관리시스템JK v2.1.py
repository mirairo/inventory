#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
- 판매재고관리시스템JK v2.1
- SQLite 데이터베이스 내장형
- sv-ttk 라이브러리를 통한 UI 개선
- 엑셀 파일 일괄 업로드 기능 추가
- 라이브러리 자동 설치 기능 포함
- 제품명, 거래처명 검색 기능 추가
"""

import subprocess
import sys
import tkinter as tk
from tkinter import messagebox, filedialog

# sv_ttk 자동 설치
try:
    import sv_ttk
except ImportError:
    print("sv_ttk 라이브러리가 설치되지 않았습니다. 자동 설치를 시도합니다.")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "sv-ttk"])
        import sv_ttk
        print("sv_ttk 라이브러리를 성공적으로 설치했습니다.")
    except Exception as e:
        messagebox.showerror("설치 오류", f"sv-ttk 설치 실패: {e}")
        sys.exit(1)

# openpyxl 자동 설치
try:
    import openpyxl
except ImportError:
    print("openpyxl 라이브러리가 설치되지 않았습니다. 자동 설치를 시도합니다.")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        print("openpyxl 라이브러리를 성공적으로 설치했습니다.")
    except Exception as e:
        messagebox.showerror("설치 오류", f"openpyxl 설치 실패: {e}")
        sys.exit(1)

import sqlite3
from tkinter import ttk, scrolledtext
from datetime import datetime
import os
import shutil


class InventoryManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("판매재고관리시스템JK v2.1")
        self.root.geometry("1280x768")

        self.db_path = self.get_db_path()
        self.conn = None
        self.cursor = None
        self.init_database()
        self.setup_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def get_db_path(self):
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        else:
            application_path = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(application_path, 'inventory.db')

    def init_database(self):
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.cursor = self.conn.cursor()

            queries = [
                '''CREATE TABLE IF NOT EXISTS products (
                    product_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_code TEXT UNIQUE NOT NULL,
                    product_name TEXT NOT NULL,
                    category TEXT,
                    unit_price REAL NOT NULL,
                    supplier TEXT,
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )''',
                '''CREATE TABLE IF NOT EXISTS inventory (
                    inventory_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER NOT NULL,
                    quantity INTEGER NOT NULL DEFAULT 0,
                    min_quantity INTEGER DEFAULT 10,
                    location TEXT,
                    last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (product_id) REFERENCES products(product_id) ON DELETE CASCADE
                )''',
                '''CREATE TABLE IF NOT EXISTS customers (
                    customer_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    customer_code TEXT UNIQUE NOT NULL,
                    customer_name TEXT NOT NULL,
                    contact TEXT,
                    address TEXT,
                    email TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )''',
                '''CREATE TABLE IF NOT EXISTS sales (
                    sale_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    customer_id INTEGER NOT NULL,
                    sale_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    total_amount REAL NOT NULL,
                    payment_status TEXT DEFAULT '미수',
                    notes TEXT,
                    FOREIGN KEY (customer_id) REFERENCES customers(customer_id) ON DELETE CASCADE
                )''',
                '''CREATE TABLE IF NOT EXISTS sale_details (
                    detail_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sale_id INTEGER NOT NULL,
                    product_id INTEGER NOT NULL,
                    quantity INTEGER NOT NULL,
                    unit_price REAL NOT NULL,
                    subtotal REAL NOT NULL,
                    FOREIGN KEY (sale_id) REFERENCES sales(sale_id) ON DELETE CASCADE,
                    FOREIGN KEY (product_id) REFERENCES products(product_id) ON DELETE CASCADE
                )''',
                '''CREATE TABLE IF NOT EXISTS transactions (
                    transaction_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER NOT NULL,
                    transaction_type TEXT NOT NULL,
                    quantity INTEGER NOT NULL,
                    transaction_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    notes TEXT,
                    FOREIGN KEY (product_id) REFERENCES products(product_id) ON DELETE CASCADE
                )'''
            ]
            for query in queries:
                self.cursor.execute(query)

            self.cursor.execute("PRAGMA foreign_keys = ON;")
            self.conn.commit()

        except sqlite3.Error as e:
            messagebox.showerror("데이터베이스 오류", f"데이터베이스 초기화 실패: {e}")
            self.root.quit()

    def setup_ui(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="파일", menu=file_menu)
        file_menu.add_command(label="데이터베이스 백업", command=self.backup_database)
        file_menu.add_separator()
        file_menu.add_command(label="종료", command=self.on_closing)

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="도움말", menu=help_menu)
        help_menu.add_command(label="사용법", command=self.show_help)
        help_menu.add_command(label="정보", command=self.show_about)

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)

        self.create_product_tab()
        self.create_inventory_tab()
        self.create_customer_tab()
        self.create_sales_tab()
        self.create_report_tab()

    def create_product_tab(self):
        tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(tab, text="  제품 관리  ")

        paned_window = ttk.PanedWindow(tab, orient='vertical')
        paned_window.pack(fill='both', expand=True)

        input_frame = ttk.LabelFrame(paned_window, text="제품 정보", padding=15)
        paned_window.add(input_frame, weight=1)

        input_frame.columnconfigure(1, weight=1)
        input_frame.columnconfigure(3, weight=1)

        ttk.Label(input_frame, text="제품 코드:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.product_code_entry = ttk.Entry(input_frame, width=25)
        self.product_code_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(input_frame, text="제품명:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.product_name_entry = ttk.Entry(input_frame, width=40)
        self.product_name_entry.grid(row=0, column=3, sticky='ew', padx=5, pady=5)

        ttk.Label(input_frame, text="카테고리:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.category_entry = ttk.Entry(input_frame, width=25)
        self.category_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(input_frame, text="단가:").grid(row=1, column=2, sticky='w', padx=5, pady=5)
        self.unit_price_entry = ttk.Entry(input_frame, width=20)
        self.unit_price_entry.grid(row=1, column=3, sticky='w', padx=5, pady=5)

        ttk.Label(input_frame, text="공급업체:").grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.supplier_entry = ttk.Entry(input_frame)
        self.supplier_entry.grid(row=2, column=1, columnspan=3, sticky='ew', padx=5, pady=5)

        ttk.Label(input_frame, text="설명:").grid(row=3, column=0, sticky='w', padx=5, pady=5)
        self.description_entry = ttk.Entry(input_frame)
        self.description_entry.grid(row=3, column=1, columnspan=3, sticky='ew', padx=5, pady=5)

        btn_frame = ttk.Frame(input_frame)
        btn_frame.grid(row=4, column=0, columnspan=4, pady=15)
        ttk.Button(btn_frame, text="등록", command=self.add_product).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="수정", command=self.update_product).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="삭제", command=self.delete_product).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="초기화", command=self.clear_product_fields).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="엑셀 첨부", command=self.import_products_from_excel).pack(side='left', padx=5)

        list_frame = ttk.LabelFrame(paned_window, text="제품 목록", padding=15)
        paned_window.add(list_frame, weight=3)
        
        # --- 제품 검색 기능 추가 ---
        search_frame = ttk.Frame(list_frame)
        search_frame.pack(fill='x', padx=5, pady=(0, 10))
        ttk.Label(search_frame, text="제품명 검색:").pack(side='left', padx=(0, 5))
        self.product_search_entry = ttk.Entry(search_frame, width=40)
        self.product_search_entry.pack(side='left', fill='x', expand=True, padx=5)
        self.product_search_entry.bind("<Return>", lambda event: self.search_products())
        ttk.Button(search_frame, text="검색", command=self.search_products).pack(side='left', padx=5)
        ttk.Button(search_frame, text="전체보기", command=lambda: self.load_products()).pack(side='left', padx=5)
        # -------------------------

        self.setup_treeview(list_frame, 'product')
        self.load_products()

    def create_inventory_tab(self):
        tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(tab, text="  재고 관리  ")

        paned_window = ttk.PanedWindow(tab, orient='vertical')
        paned_window.pack(fill='both', expand=True)

        trans_frame = ttk.LabelFrame(paned_window, text="입/출고 처리", padding=15)
        paned_window.add(trans_frame, weight=1)
        trans_frame.columnconfigure(1, weight=1)

        ttk.Label(trans_frame, text="제품:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.inv_product_combo = ttk.Combobox(trans_frame, width=40, state='readonly')
        self.inv_product_combo.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(trans_frame, text="수량:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.inv_quantity_entry = ttk.Entry(trans_frame, width=15)
        self.inv_quantity_entry.grid(row=0, column=3, sticky='w', padx=5, pady=5)

        ttk.Label(trans_frame, text="비고:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.inv_notes_entry = ttk.Entry(trans_frame)
        self.inv_notes_entry.grid(row=1, column=1, columnspan=3, sticky='ew', padx=5, pady=5)

        btn_frame = ttk.Frame(trans_frame)
        btn_frame.grid(row=2, column=0, columnspan=4, pady=15)
        ttk.Button(btn_frame, text="입고 처리", command=lambda: self.process_transaction('입고')).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="출고 처리", command=lambda: self.process_transaction('출고')).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="새로고침", command=self.load_inventory).pack(side='left', padx=5)

        inv_frame = ttk.LabelFrame(paned_window, text="재고 현황", padding=15)
        paned_window.add(inv_frame, weight=3)
        self.setup_treeview(inv_frame, 'inventory')
        self.load_product_combo()
        self.load_inventory()

    def create_customer_tab(self):
        tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(tab, text="  거래처 관리  ")

        paned_window = ttk.PanedWindow(tab, orient='vertical')
        paned_window.pack(fill='both', expand=True)

        input_frame = ttk.LabelFrame(paned_window, text="거래처 정보", padding=15)
        paned_window.add(input_frame, weight=1)
        input_frame.columnconfigure(1, weight=1)
        input_frame.columnconfigure(3, weight=1)

        ttk.Label(input_frame, text="거래처 코드:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.customer_code_entry = ttk.Entry(input_frame, width=25)
        self.customer_code_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(input_frame, text="거래처명:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.customer_name_entry = ttk.Entry(input_frame, width=40)
        self.customer_name_entry.grid(row=0, column=3, sticky='ew', padx=5, pady=5)
        
        ttk.Label(input_frame, text="연락처:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.contact_entry = ttk.Entry(input_frame, width=25)
        self.contact_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=5)
        
        ttk.Label(input_frame, text="이메일:").grid(row=1, column=2, sticky='w', padx=5, pady=5)
        self.email_entry = ttk.Entry(input_frame, width=40)
        self.email_entry.grid(row=1, column=3, sticky='ew', padx=5, pady=5)
        
        ttk.Label(input_frame, text="주소:").grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.address_entry = ttk.Entry(input_frame)
        self.address_entry.grid(row=2, column=1, columnspan=3, sticky='ew', padx=5, pady=5)
        
        btn_frame = ttk.Frame(input_frame)
        btn_frame.grid(row=3, column=0, columnspan=4, pady=15)
        ttk.Button(btn_frame, text="등록", command=self.add_customer).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="수정", command=self.update_customer).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="삭제", command=self.delete_customer).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="초기화", command=self.clear_customer_fields).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="엑셀 첨부", command=self.import_customers_from_excel).pack(side='left', padx=5)

        list_frame = ttk.LabelFrame(paned_window, text="거래처 목록", padding=15)
        paned_window.add(list_frame, weight=3)
        
        # --- 거래처 검색 기능 추가 ---
        search_frame = ttk.Frame(list_frame)
        search_frame.pack(fill='x', padx=5, pady=(0, 10))
        ttk.Label(search_frame, text="거래처명 검색:").pack(side='left', padx=(0, 5))
        self.customer_search_entry = ttk.Entry(search_frame, width=40)
        self.customer_search_entry.pack(side='left', fill='x', expand=True, padx=5)
        self.customer_search_entry.bind("<Return>", lambda event: self.search_customers())
        ttk.Button(search_frame, text="검색", command=self.search_customers).pack(side='left', padx=5)
        ttk.Button(search_frame, text="전체보기", command=lambda: self.load_customers()).pack(side='left', padx=5)
        # ---------------------------

        self.setup_treeview(list_frame, 'customer')
        self.load_customers()

    def create_sales_tab(self):
        tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(tab, text="  판매 관리  ")

        paned_window = ttk.PanedWindow(tab, orient='vertical')
        paned_window.pack(fill='both', expand=True)

        input_frame = ttk.LabelFrame(paned_window, text="판매 정보 입력", padding=15)
        paned_window.add(input_frame, weight=1)
        input_frame.columnconfigure(1, weight=1)
        input_frame.columnconfigure(3, weight=1)

        ttk.Label(input_frame, text="거래처:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.sale_customer_combo = ttk.Combobox(input_frame, width=30, state='readonly')
        self.sale_customer_combo.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(input_frame, text="제품:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.sale_product_combo = ttk.Combobox(input_frame, width=30, state='readonly')
        self.sale_product_combo.grid(row=0, column=3, sticky='ew', padx=5, pady=5)
        self.sale_product_combo.bind("<<ComboboxSelected>>", self.on_sale_product_select)

        ttk.Label(input_frame, text="수량:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.sale_quantity_entry = ttk.Entry(input_frame, width=15)
        self.sale_quantity_entry.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(input_frame, text="단가:").grid(row=1, column=2, sticky='w', padx=5, pady=5)
        self.sale_price_entry = ttk.Entry(input_frame, width=15)
        self.sale_price_entry.grid(row=1, column=3, sticky='w', padx=5, pady=5)
        
        ttk.Label(input_frame, text="비고:").grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.sale_notes_entry = ttk.Entry(input_frame)
        self.sale_notes_entry.grid(row=2, column=1, columnspan=3, sticky='ew', padx=5, pady=5)
        
        btn_frame = ttk.Frame(input_frame)
        btn_frame.grid(row=3, column=0, columnspan=4, pady=15)
        ttk.Button(btn_frame, text="판매 등록", command=self.add_sale).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="판매 내역 삭제", command=self.delete_sale).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="새로고침", command=self.load_sales).pack(side='left', padx=5)

        list_frame = ttk.LabelFrame(paned_window, text="판매 내역", padding=15)
        paned_window.add(list_frame, weight=3)
        self.setup_treeview(list_frame, 'sales')
        self.load_customer_combo()
        self.load_sale_product_combo()
        self.load_sales()

    def create_report_tab(self):
        tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(tab, text="  통계 및 보고서  ")

        top_frame = ttk.Frame(tab, padding=5)
        top_frame.pack(fill='x')
        ttk.Button(top_frame, text="재고 부족 제품", command=self.show_low_stock).pack(side='left', padx=5, pady=5)
        ttk.Button(top_frame, text="월별 매출 통계", command=self.show_monthly_sales).pack(side='left', padx=5, pady=5)
        ttk.Button(top_frame, text="입/출고 내역", command=self.show_transactions).pack(side='left', padx=5, pady=5)

        report_frame = ttk.LabelFrame(tab, text="보고서", padding=15)
        report_frame.pack(fill='both', expand=True, padx=5, pady=5)
        self.report_text = scrolledtext.ScrolledText(report_frame, width=100, height=30, font=('맑은 고딕', 10), relief='flat')
        self.report_text.pack(fill='both', expand=True)
        self.report_text.config(state='disabled')

    def setup_treeview(self, parent, view_type):
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill='both', expand=True)
        
        columns = ()
        if view_type == 'product':
            columns = ('제품코드', '제품명', '카테고리', '단가', '공급업체')
        elif view_type == 'inventory':
            columns = ('제품코드', '제품명', '현재고', '최소재고', '위치', '최종수정일')
        elif view_type == 'customer':
            columns = ('거래처코드', '거래처명', '연락처', '이메일', '주소')
        elif view_type == 'sales':
            columns = ('판매일', '거래처', '제품명', '수량', '단가', '합계', '상태')
        
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        
        tree.heading(columns[0], text=columns[0])
        tree.column(columns[0], width=100, anchor='center')
        for col in columns[1:]:
            tree.heading(col, text=col)
            tree.column(col, width=150)
        
        scrollbar_y = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        scrollbar_y.pack(side='right', fill='y')
        scrollbar_x.pack(side='bottom', fill='x')
        tree.pack(side='left', fill='both', expand=True)

        if view_type == 'product':
            self.product_tree = tree
            tree.bind('<Double-1>', self.on_product_select)
        elif view_type == 'inventory':
            self.inventory_tree = tree
        elif view_type == 'customer':
            self.customer_tree = tree
            tree.bind('<Double-1>', self.on_customer_select)
        elif view_type == 'sales':
            self.sales_tree = tree

    def import_products_from_excel(self):
        file_path = filedialog.askopenfilename(
            title="제품 목록 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            success_count = 0
            error_count = 0
            error_messages = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0] or not row[1]:
                    continue
                
                try:
                    product_code = str(row[0]).strip()
                    product_name = str(row[1]).strip()
                    category = str(row[2]).strip() if row[2] else ""
                    unit_price = float(row[3]) if row[3] else 0
                    supplier = str(row[4]).strip() if row[4] else ""
                    description = str(row[5]).strip() if row[5] else ""
                    
                    self.cursor.execute(
                        'INSERT INTO products (product_code, product_name, category, unit_price, supplier, description) VALUES (?, ?, ?, ?, ?, ?)',
                        (product_code, product_name, category, unit_price, supplier, description)
                    )
                    product_id = self.cursor.lastrowid
                    self.cursor.execute('INSERT INTO inventory (product_id) VALUES (?)', (product_id,))
                    success_count += 1
                    
                except sqlite3.IntegrityError:
                    error_count += 1
                    error_messages.append(f"행 {row_idx}: 중복된 제품 코드 ({product_code})")
                except (ValueError, TypeError) as e:
                    error_count += 1
                    error_messages.append(f"행 {row_idx}: 데이터 형식 오류 ({str(e)})")
                except Exception as e:
                    error_count += 1
                    error_messages.append(f"행 {row_idx}: {str(e)}")
            
            self.conn.commit()
            wb.close()
            
            result_msg = f"등록 완료: {success_count}건\n실패: {error_count}건"
            if error_messages:
                result_msg += "\n\n[오류 상세]:\n" + "\n".join(error_messages[:10])
                if len(error_messages) > 10:
                    result_msg += f"\n... 외 {len(error_messages) - 10}건"
            
            messagebox.showinfo("엑셀 가져오기 완료", result_msg)
            self.load_products()
            self.load_product_combo()
            self.load_sale_product_combo()
            
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 파일 읽기 실패: {e}")

    def import_customers_from_excel(self):
        file_path = filedialog.askopenfilename(
            title="거래처 목록 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            success_count = 0
            error_count = 0
            error_messages = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0] or not row[1]:
                    continue
                
                try:
                    customer_code = str(row[0]).strip()
                    customer_name = str(row[1]).strip()
                    contact = str(row[2]).strip() if row[2] else ""
                    email = str(row[3]).strip() if row[3] else ""
                    address = str(row[4]).strip() if row[4] else ""
                    
                    self.cursor.execute(
                        'INSERT INTO customers (customer_code, customer_name, contact, email, address) VALUES (?, ?, ?, ?, ?)',
                        (customer_code, customer_name, contact, email, address)
                    )
                    success_count += 1
                    
                except sqlite3.IntegrityError:
                    error_count += 1
                    error_messages.append(f"행 {row_idx}: 중복된 거래처 코드 ({customer_code})")
                except (ValueError, TypeError) as e:
                    error_count += 1
                    error_messages.append(f"행 {row_idx}: 데이터 형식 오류 ({str(e)})")
                except Exception as e:
                    error_count += 1
                    error_messages.append(f"행 {row_idx}: {str(e)}")
            
            self.conn.commit()
            wb.close()
            
            result_msg = f"등록 완료: {success_count}건\n실패: {error_count}건"
            if error_messages:
                result_msg += "\n\n[오류 상세]:\n" + "\n".join(error_messages[:10])
                if len(error_messages) > 10:
                    result_msg += f"\n... 외 {len(error_messages) - 10}건"
            
            messagebox.showinfo("엑셀 가져오기 완료", result_msg)
            self.load_customers()
            self.load_customer_combo()
            
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 파일 읽기 실패: {e}")

    def add_product(self):
        code = self.product_code_entry.get().strip()
        name = self.product_name_entry.get().strip()
        price = self.unit_price_entry.get().strip()
        if not code or not name or not price:
            messagebox.showwarning("입력 오류", "제품 코드, 제품명, 단가는 필수 항목입니다.")
            return
        try:
            price_val = float(price)
            self.cursor.execute('INSERT INTO products (product_code, product_name, category, unit_price, supplier, description) VALUES (?, ?, ?, ?, ?, ?)',
                                (code, name, self.category_entry.get().strip(), price_val, self.supplier_entry.get().strip(), self.description_entry.get().strip()))
            product_id = self.cursor.lastrowid
            self.cursor.execute('INSERT INTO inventory (product_id) VALUES (?)', (product_id,))
            self.conn.commit()
            messagebox.showinfo("성공", "제품이 등록되었습니다.")
            self.clear_product_fields()
            self.load_products()
            self.load_product_combo()
            self.load_sale_product_combo()
        except sqlite3.IntegrityError:
            messagebox.showerror("오류", "이미 존재하는 제품 코드입니다.")
        except ValueError:
            messagebox.showerror("입력 오류", "단가는 숫자로 입력해야 합니다.")
        except Exception as e:
            messagebox.showerror("오류", f"제품 등록 중 오류 발생: {e}")

    def update_product(self):
        selected = self.product_tree.selection()
        if not selected:
            messagebox.showwarning("선택 오류", "수정할 제품을 선택해주세요.")
            return
        try:
            item = self.product_tree.item(selected[0])
            product_code_from_tree = item['values'][0]
            self.cursor.execute('SELECT product_id FROM products WHERE product_code=?', (product_code_from_tree,))
            result = self.cursor.fetchone()
            if not result:
                messagebox.showerror("오류", "선택된 제품을 DB에서 찾을 수 없습니다.")
                return
            product_id = result[0]
            code = self.product_code_entry.get().strip()
            price = float(self.unit_price_entry.get().strip())
            self.cursor.execute('UPDATE products SET product_code=?, product_name=?, category=?, unit_price=?, supplier=?, description=? WHERE product_id=?',
                                (code, self.product_name_entry.get().strip(), self.category_entry.get().strip(), price, self.supplier_entry.get().strip(), self.description_entry.get().strip(), product_id))
            self.conn.commit()
            messagebox.showinfo("성공", "제품 정보가 수정되었습니다.")
            self.clear_product_fields()
            self.load_products()
        except ValueError:
            messagebox.showerror("입력 오류", "단가는 숫자로 입력해야 합니다.")
        except Exception as e:
            messagebox.showerror("오류", f"제품 수정 중 오류 발생: {e}")

    def delete_product(self):
        selected = self.product_tree.selection()
        if not selected:
            messagebox.showwarning("선택 오류", "삭제할 제품을 선택해주세요.")
            return
        if not messagebox.askyesno("삭제 확인", "선택한 제품을 삭제하시겠습니까?\n연관된 재고, 판매 내역도 모두 삭제됩니다."):
            return
        try:
            item = self.product_tree.item(selected[0])
            product_code = item['values'][0]
            self.cursor.execute('DELETE FROM products WHERE product_code=?', (product_code,))
            self.conn.commit()
            messagebox.showinfo("성공", "제품이 삭제되었습니다.")
            self.clear_product_fields()
            self.load_products()
            self.load_inventory()
            self.load_sales()
        except Exception as e:
            messagebox.showerror("오류", f"제품 삭제 중 오류 발생: {e}")

    def clear_product_fields(self):
        self.product_code_entry.delete(0, tk.END)
        self.product_name_entry.delete(0, tk.END)
        self.category_entry.delete(0, tk.END)
        self.unit_price_entry.delete(0, tk.END)
        self.supplier_entry.delete(0, tk.END)
        self.description_entry.delete(0, tk.END)
        if self.product_tree.selection():
            self.product_tree.selection_remove(self.product_tree.selection()[0])

    def load_products(self, search_term=None):
        for item in self.product_tree.get_children():
            self.product_tree.delete(item)
        
        if search_term:
            query = 'SELECT product_code, product_name, category, unit_price, supplier FROM products WHERE product_name LIKE ? ORDER BY product_id DESC'
            params = (f'%{search_term}%',)
        else:
            query = 'SELECT product_code, product_name, category, unit_price, supplier FROM products ORDER BY product_id DESC'
            params = ()
            self.product_search_entry.delete(0, tk.END)

        self.cursor.execute(query, params)
        for row in self.cursor.fetchall():
            self.product_tree.insert('', 'end', values=(row[0], row[1], row[2], f"{row[3]:,.0f}", row[4]))

    def search_products(self):
        search_term = self.product_search_entry.get().strip()
        if not search_term:
            messagebox.showwarning("검색 오류", "검색할 제품명을 입력하세요.")
            return
        self.load_products(search_term)

    def on_product_select(self, event):
        selected = self.product_tree.selection()
        if selected:
            item = self.product_tree.item(selected[0])['values']
            product_code = item[0]
            self.cursor.execute('SELECT * FROM products WHERE product_code=?', (product_code,))
            row = self.cursor.fetchone()
            self.clear_product_fields()
            self.product_code_entry.insert(0, row[1])
            self.product_name_entry.insert(0, row[2])
            self.category_entry.insert(0, row[3] or '')
            self.unit_price_entry.insert(0, str(row[4]))
            self.supplier_entry.insert(0, row[5] or '')
            self.description_entry.insert(0, row[6] or '')

    def load_product_combo(self):
        self.cursor.execute('SELECT product_code, product_name FROM products ORDER BY product_name')
        products = [f"{row[0]} - {row[1]}" for row in self.cursor.fetchall()]
        self.inv_product_combo['values'] = products

    def process_transaction(self, trans_type):
        product_str = self.inv_product_combo.get()
        quantity_str = self.inv_quantity_entry.get().strip()
        if not product_str or not quantity_str:
            messagebox.showwarning("입력 오류", "제품과 수량을 모두 입력해주세요.")
            return
        try:
            product_code = product_str.split(' - ')[0]
            quantity = int(quantity_str)
            if quantity <= 0:
                messagebox.showerror("입력 오류", "수량은 0보다 커야 합니다.")
                return
            self.cursor.execute('SELECT p.product_id, i.quantity FROM products p JOIN inventory i ON p.product_id = i.product_id WHERE p.product_code=?', (product_code,))
            result = self.cursor.fetchone()
            if not result:
                messagebox.showerror("오류", "제품을 찾을 수 없습니다.")
                return
            product_id, current_qty = result
            if trans_type == '출고' and current_qty < quantity:
                messagebox.showerror("재고 부족", f"현재 재고 ({current_qty})가 부족하여 출고할 수 없습니다.")
                return
            new_qty = current_qty + quantity if trans_type == '입고' else current_qty - quantity
            self.cursor.execute('UPDATE inventory SET quantity=?, last_updated=CURRENT_TIMESTAMP WHERE product_id=?', (new_qty, product_id))
            self.cursor.execute('INSERT INTO transactions (product_id, transaction_type, quantity, notes) VALUES (?, ?, ?, ?)',
                                (product_id, trans_type, quantity, self.inv_notes_entry.get().strip()))
            self.conn.commit()
            messagebox.showinfo("성공", f"{trans_type} 처리가 완료되었습니다.")
            self.inv_quantity_entry.delete(0, tk.END)
            self.inv_notes_entry.delete(0, tk.END)
            self.inv_product_combo.set('')
            self.load_inventory()
        except ValueError:
            messagebox.showerror("입력 오류", "수량은 숫자로 입력해주세요.")
        except Exception as e:
            messagebox.showerror("오류", f"처리 중 오류 발생: {e}")

    def load_inventory(self):
        for item in self.inventory_tree.get_children():
            self.inventory_tree.delete(item)
        self.inventory_tree.tag_configure('low_stock', background='#FFCDD2', foreground='#B71C1C')
        query = '''
            SELECT p.product_code, p.product_name, i.quantity, i.min_quantity, i.location, i.last_updated
            FROM inventory i JOIN products p ON i.product_id = p.product_id
            ORDER BY p.product_name
        '''
        self.cursor.execute(query)
        for row in self.cursor.fetchall():
            tags = ()
            if row[2] <= row[3]:
                tags = ('low_stock',)
            values = (row[0], row[1], row[2], row[3], row[4] or '', row[5][:16] if row[5] else '')
            self.inventory_tree.insert('', 'end', values=values, tags=tags)

    def add_customer(self):
        code = self.customer_code_entry.get().strip()
        name = self.customer_name_entry.get().strip()
        if not code or not name:
            messagebox.showwarning("입력 오류", "거래처 코드와 거래처명은 필수입니다.")
            return
        try:
            self.cursor.execute('INSERT INTO customers (customer_code, customer_name, contact, email, address) VALUES (?, ?, ?, ?, ?)',
                                (code, name, self.contact_entry.get().strip(), self.email_entry.get().strip(), self.address_entry.get().strip()))
            self.conn.commit()
            messagebox.showinfo("성공", "거래처가 등록되었습니다.")
            self.clear_customer_fields()
            self.load_customers()
            self.load_customer_combo()
        except sqlite3.IntegrityError:
            messagebox.showerror("오류", "이미 존재하는 거래처 코드입니다.")
        except Exception as e:
            messagebox.showerror("오류", f"거래처 등록 중 오류: {e}")

    def update_customer(self):
        selected = self.customer_tree.selection()
        if not selected:
            messagebox.showwarning("선택 오류", "수정할 거래처를 선택해주세요.")
            return
        try:
            item = self.customer_tree.item(selected[0])['values']
            customer_code_from_tree = item[0]
            self.cursor.execute('SELECT customer_id FROM customers WHERE customer_code=?', (customer_code_from_tree,))
            customer_id = self.cursor.fetchone()[0]
            code = self.customer_code_entry.get().strip()
            name = self.customer_name_entry.get().strip()
            self.cursor.execute('UPDATE customers SET customer_code=?, customer_name=?, contact=?, email=?, address=? WHERE customer_id=?',
                                (code, name, self.contact_entry.get().strip(), self.email_entry.get().strip(), self.address_entry.get().strip(), customer_id))
            self.conn.commit()
            messagebox.showinfo("성공", "거래처 정보가 수정되었습니다.")
            self.clear_customer_fields()
            self.load_customers()
        except Exception as e:
            messagebox.showerror("오류", f"거래처 수정 중 오류: {e}")

    def delete_customer(self):
        selected = self.customer_tree.selection()
        if not selected:
            messagebox.showwarning("선택 오류", "삭제할 거래처를 선택해주세요.")
            return
        if not messagebox.askyesno("삭제 확인", "선택한 거래처를 삭제하시겠습니까?\n연관된 판매 내역도 모두 삭제됩니다."):
            return
        try:
            item = self.customer_tree.item(selected[0])['values']
            customer_code = item[0]
            self.cursor.execute('DELETE FROM customers WHERE customer_code=?', (customer_code,))
            self.conn.commit()
            messagebox.showinfo("성공", "거래처가 삭제되었습니다.")
            self.clear_customer_fields()
            self.load_customers()
            self.load_sales()
        except Exception as e:
            messagebox.showerror("오류", f"거래처 삭제 중 오류: {e}")

    def clear_customer_fields(self):
        self.customer_code_entry.delete(0, tk.END)
        self.customer_name_entry.delete(0, tk.END)
        self.contact_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
        self.address_entry.delete(0, tk.END)
        if self.customer_tree.selection():
            self.customer_tree.selection_remove(self.customer_tree.selection()[0])

    def load_customers(self, search_term=None):
        for item in self.customer_tree.get_children():
            self.customer_tree.delete(item)
            
        if search_term:
            query = 'SELECT customer_code, customer_name, contact, email, address FROM customers WHERE customer_name LIKE ? ORDER BY customer_id DESC'
            params = (f'%{search_term}%',)
        else:
            query = 'SELECT customer_code, customer_name, contact, email, address FROM customers ORDER BY customer_id DESC'
            params = ()
            self.customer_search_entry.delete(0, tk.END)

        self.cursor.execute(query, params)
        for row in self.cursor.fetchall():
            self.customer_tree.insert('', 'end', values=row)

    def search_customers(self):
        search_term = self.customer_search_entry.get().strip()
        if not search_term:
            messagebox.showwarning("검색 오류", "검색할 거래처명을 입력하세요.")
            return
        self.load_customers(search_term)

    def on_customer_select(self, event):
        selected = self.customer_tree.selection()
        if selected:
            item = self.customer_tree.item(selected[0])['values']
            self.clear_customer_fields()
            self.customer_code_entry.insert(0, item[0])
            self.customer_name_entry.insert(0, item[1])
            self.contact_entry.insert(0, item[2])
            self.email_entry.insert(0, item[3])
            self.address_entry.insert(0, item[4])

    def load_customer_combo(self):
        self.cursor.execute('SELECT customer_code, customer_name FROM customers ORDER BY customer_name')
        customers = [f"{row[0]} - {row[1]}" for row in self.cursor.fetchall()]
        self.sale_customer_combo['values'] = customers

    def load_sale_product_combo(self):
        self.cursor.execute('SELECT product_code, product_name FROM products ORDER BY product_name')
        products = [f"{row[0]} - {row[1]}" for row in self.cursor.fetchall()]
        self.sale_product_combo['values'] = products

    def on_sale_product_select(self, event):
        product_str = self.sale_product_combo.get()
        if not product_str:
            return
        product_code = product_str.split(' - ')[0]
        self.cursor.execute('SELECT unit_price FROM products WHERE product_code=?', (product_code,))
        result = self.cursor.fetchone()
        if result:
            self.sale_price_entry.delete(0, tk.END)
            self.sale_price_entry.insert(0, str(result[0]))

    def add_sale(self):
        customer_str = self.sale_customer_combo.get()
        product_str = self.sale_product_combo.get()
        qty_str = self.sale_quantity_entry.get().strip()
        price_str = self.sale_price_entry.get().strip()
        if not all([customer_str, product_str, qty_str, price_str]):
            messagebox.showwarning("입력 오류", "모든 판매 정보를 입력해주세요.")
            return
        try:
            customer_code = customer_str.split(' - ')[0]
            product_code = product_str.split(' - ')[0]
            quantity = int(qty_str)
            price = float(price_str)
            subtotal = quantity * price
            self.cursor.execute('SELECT customer_id FROM customers WHERE customer_code=?', (customer_code,))
            customer_id = self.cursor.fetchone()[0]
            self.cursor.execute('SELECT product_id FROM products WHERE product_code=?', (product_code,))
            product_id = self.cursor.fetchone()[0]
            self.cursor.execute('SELECT quantity FROM inventory WHERE product_id=?', (product_id,))
            current_qty = self.cursor.fetchone()[0]
            if current_qty < quantity:
                if not messagebox.askyesno("재고 부족", f"현재고 ({current_qty})가 부족합니다. 마이너스 재고로 판매하시겠습니까?"):
                    return
            self.cursor.execute('INSERT INTO sales (customer_id, total_amount, notes) VALUES (?, ?, ?)', (customer_id, subtotal, self.sale_notes_entry.get().strip()))
            sale_id = self.cursor.lastrowid
            self.cursor.execute('INSERT INTO sale_details (sale_id, product_id, quantity, unit_price, subtotal) VALUES (?, ?, ?, ?, ?)', (sale_id, product_id, quantity, price, subtotal))
            new_qty = current_qty - quantity
            self.cursor.execute('UPDATE inventory SET quantity=? WHERE product_id=?', (new_qty, product_id))
            self.cursor.execute("INSERT INTO transactions (product_id, transaction_type, quantity, notes) VALUES (?, '판매', ?, ?)", (product_id, quantity, f"판매번호: {sale_id}"))
            self.conn.commit()
            messagebox.showinfo("성공", f"판매가 등록되었습니다.\n판매금액: {subtotal:,.0f}원")
            self.load_sales()
            self.load_inventory()
        except (ValueError, TypeError):
            messagebox.showerror("입력 오류", "수량과 단가는 숫자로 입력해야 합니다.")
        except Exception as e:
            messagebox.showerror("오류", f"판매 등록 중 오류: {e}")
            self.conn.rollback()

    def delete_sale(self):
        selected = self.sales_tree.selection()
        if not selected:
            messagebox.showwarning("선택 오류", "삭제할 판매 내역을 선택해주세요.")
            return
        if not messagebox.askyesno("삭제 확인", "선택한 판매 내역을 삭제하시겠습니까?\n차감되었던 재고가 다시 복구됩니다."):
            return
        try:
            item = self.sales_tree.item(selected[0])
            sale_id = item['text']
            self.cursor.execute('SELECT product_id, quantity FROM sale_details WHERE sale_id=?', (sale_id,))
            detail = self.cursor.fetchone()
            if not detail:
                messagebox.showerror("오류", "판매 상세 내역을 찾을 수 없습니다.")
                return
            product_id, quantity = detail
            self.cursor.execute('UPDATE inventory SET quantity = quantity + ? WHERE product_id=?', (quantity, product_id))
            self.cursor.execute('DELETE FROM sales WHERE sale_id=?', (sale_id,))
            self.cursor.execute("INSERT INTO transactions (product_id, transaction_type, quantity, notes) VALUES (?, '판매취소', ?, ?)", (product_id, quantity, f"취소된 판매번호: {sale_id}"))
            self.conn.commit()
            messagebox.showinfo("성공", "판매 내역이 삭제되고 재고가 복구되었습니다.")
            self.load_sales()
            self.load_inventory()
        except Exception as e:
            messagebox.showerror("오류", f"판매 내역 삭제 중 오류 발생: {e}")
            self.conn.rollback()

    def load_sales(self):
        for item in self.sales_tree.get_children():
            self.sales_tree.delete(item)
        query = '''
            SELECT s.sale_id, s.sale_date, c.customer_name, p.product_name, sd.quantity, sd.unit_price, sd.subtotal, s.payment_status
            FROM sales s
            JOIN customers c ON s.customer_id = c.customer_id
            JOIN sale_details sd ON s.sale_id = sd.sale_id
            JOIN products p ON sd.product_id = p.product_id
            ORDER BY s.sale_id DESC LIMIT 200
        '''
        self.cursor.execute(query)
        for row in self.cursor.fetchall():
            values = (row[1][:16], row[2], row[3], row[4], f"{row[5]:,.0f}", f"{row[6]:,.0f}", row[7])
            self.sales_tree.insert('', 'end', text=row[0], values=values)

    def _generate_report(self, title, headers, data_rows):
        self.report_text.config(state='normal')
        self.report_text.delete(1.0, tk.END)
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        report = f"{'='*80}\n"
        report += f"{title:^80}\n"
        report += f"보고서 생성일시: {now}\n"
        report += f"{'='*80}\n\n"
        if not data_rows:
            report += "표시할 데이터가 없습니다.\n"
        else:
            header_str = "".join([f"{h:<{w}}" for h, w in headers])
            report += f"{header_str}\n"
            report += f"{'-'*80}\n"
            for row in data_rows:
                row_str = "".join([f"{str(item):<{w}}" for item, (h, w) in zip(row, headers)])
                report += f"{row_str}\n"
        self.report_text.insert(tk.END, report)
        self.report_text.config(state='disabled')

    def show_low_stock(self):
        self.cursor.execute('''
            SELECT p.product_code, p.product_name, i.quantity, i.min_quantity, (i.min_quantity - i.quantity) as shortage
            FROM inventory i JOIN products p ON i.product_id = p.product_id
            WHERE i.quantity <= i.min_quantity ORDER BY shortage DESC
        ''')
        headers = [('제품코드', 15), ('제품명', 30), ('현재고', 10), ('최소재고', 10), ('부족수량', 10)]
        self._generate_report("재고 부족 제품 현황", headers, self.cursor.fetchall())

    def show_monthly_sales(self):
        self.cursor.execute('''
            SELECT strftime('%Y-%m', sale_date) as month, COUNT(*), SUM(total_amount)
            FROM sales GROUP BY month ORDER BY month DESC LIMIT 12
        ''')
        results = self.cursor.fetchall()
        data_rows = []
        total_sales = 0
        for row in results:
            total_sales += row[2]
            data_rows.append((row[0], row[1], f"{row[2]:,.0f}"))
        if data_rows:
            data_rows.append(('', '', ''))
            data_rows.append(('이계', '', f"{total_sales:,.0f}"))
        headers = [('판매월', 20), ('판매건수', 20), ('매출액 (원)', 20)]
        self._generate_report("월별 매출 통계", headers, data_rows)

    def show_transactions(self):
        self.report_text.config(state='normal')
        self.report_text.delete(1.0, tk.END)
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        report = f"{'='*80}\n"
        report += f"{'입/출고 내역 (최근 100건)':^80}\n"
        report += f"보고서 생성일시: {now}\n"
        report += f"{'='*80}\n\n"
        self.cursor.execute('''
            SELECT t.transaction_date, p.product_code, p.product_name, t.transaction_type, t.quantity, t.notes
            FROM transactions t JOIN products p ON t.product_id = p.product_id
            ORDER BY t.transaction_id DESC LIMIT 100
        ''')
        results = self.cursor.fetchall()
        if not results:
            report += "입/출고 내역이 없습니다.\n"
        else:
            for row in results:
                report += f"일시: {row[0][:19]}\n"
                report += f"제품: [{row[1]}] {row[2]}\n"
                report += f"구분: {row[3]:<10} 수량: {row[4]}\n"
                if row[5]:
                    report += f"비고: {row[5]}\n"
                report += f"{'-'*80}\n"
        self.report_text.insert(tk.END, report)
        self.report_text.config(state='disabled')

    def backup_database(self):
        try:
            backup_dir = os.path.join(os.path.dirname(self.db_path), 'backup')
            os.makedirs(backup_dir, exist_ok=True)
            backup_filename = f'inventory_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
            backup_path = os.path.join(backup_dir, backup_filename)
            shutil.copy2(self.db_path, backup_path)
            messagebox.showinfo("백업 완료", f"데이터베이스가 백업되었습니다.\n경로: {backup_path}")
        except Exception as e:
            messagebox.showerror("백업 오류", f"백업 중 오류 발생: {e}")

    def show_help(self):
        help_text = """[판매재고관리시스템JK v2.1 사용법]

1. 제품 관리
   - 제품 정보를 등록, 수정, 삭제합니다
   - '제품명 검색'으로 원하는 제품을 쉽게 찾을 수 있습니다.
   - 엑셀 첨부: 엑셀 파일(.xlsx)로 제품 일괄 등록
   - 엑셀 형식: 제품코드 | 제품명 | 카테고리 | 단가 | 공급업체 | 설명

2. 재고 관리
   - 제품을 선택하여 입고/출고를 처리하고 현재고를 관리
   - 최소재고 이하인 품목은 목록에서 강조 표시

3. 거래처 관리
   - 판매 관리에 필요한 거래처 정보를 관리
   - '거래처명 검색'으로 원하는 거래처를 쉽게 찾을 수 있습니다.
   - 엑셀 첨부: 엑셀 파일(.xlsx)로 거래처 일괄 등록
   - 엑셀 형식: 거래처코드 | 거래처명 | 연락처 | 이메일 | 주소

4. 판매 관리
   - 거래처와 제품을 선택하여 판매 내역을 등록
   - 판매 등록 시 자동으로 재고가 차감됩니다

5. 통계 및 보고서
   - 재고 부족 현황, 월별 매출, 입/출고 내역 등
   - 다양한 보고서를 조회할 수 있습니다

* 데이터베이스 파일(inventory.db)은 프로그램과 동일한 폴더에 생성
* 파일 메뉴에서 데이터베이스를 안전하게 백업 가능
* 엑셀 파일 첫 번째 행은 헤더로 간주되어 자동으로 건너뜀
        """
        messagebox.showinfo("사용법", help_text)

    def show_about(self):
        about_text = """판매재고관리시스템JK v2.1

중소기업 맞춤형 판매재고관리 솔루션
개발: Python, SQLite, Tkinter, sv-ttk, openpyxl

주요 기능:
- 제품/재고/거래처/판매 통합 관리
- 제품 및 거래처 검색 기능
- 엑셀 파일 일괄 업로드 지원
- 실시간 재고 추적 및 통계 보고서
- 데이터베이스 백업 기능

© JK이러닝연구소 Jong-Ki Lee 2025 All Rights Reserved.
- [중요 안내] 이 소프트웨어는 개인 용도로만 사용하고, "배포, 전송, 공유" 등을 할 때에는, 저작권법에 따라 JK이러닝연구소로부터 문서로 허락을 받으시기 바랍니다. email: mirae91@gmail.com
        """
        messagebox.showinfo("프로그램 정보", about_text)

    def on_closing(self):
        if messagebox.askokcancel("종료", "프로그램을 종료하시겠습니까?"):
            if self.conn:
                self.conn.close()
            self.root.destroy()


def main():
    root = tk.Tk()
    app = InventoryManagementSystem(root)
    sv_ttk.set_theme("light")
    root.mainloop()


if __name__ == "__main__":
    main()
